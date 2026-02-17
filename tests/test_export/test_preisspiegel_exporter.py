from decimal import Decimal

import pytest
from openpyxl import load_workbook

from lvgenerator.export.preisspiegel_exporter import PreisSpiegelExporter
from lvgenerator.models.preisspiegel import (
    BidderInfo,
    PreisSpiegel,
    PreisSpiegelCategoryRow,
    PreisSpiegelRow,
)


def _spiegel(bidders=None, rows=None, grand_totals=None):
    return PreisSpiegel(
        project_name="Testprojekt",
        bidders=bidders or [],
        rows=rows or [],
        grand_totals=grand_totals or [],
    )


def _bidder(name, path=""):
    return BidderInfo(name=name, file_path=path)


def _item_row(oz="01.0010", text="Beton", qty=Decimal("10"), qu="m2",
              unit_prices=None, total_prices=None, not_offered=None,
              min_up=None, max_up=None, avg_up=None):
    n = len(unit_prices) if unit_prices else 0
    return PreisSpiegelRow(
        oz=oz, short_text=text, qty=qty, qu=qu,
        unit_prices=unit_prices or [],
        total_prices=total_prices or [],
        not_offered=not_offered or [False] * n,
        min_up=min_up, max_up=max_up, avg_up=avg_up,
    )


def _cat_row(oz="01", label="Rohbau", totals=None):
    return PreisSpiegelCategoryRow(oz=oz, label=label, totals=totals or [])


@pytest.fixture
def exporter():
    return PreisSpiegelExporter()


@pytest.fixture
def tmp_xlsx(tmp_path):
    return str(tmp_path / "preisspiegel.xlsx")


class TestPreisSpiegelExporter:
    def test_creates_file(self, exporter, tmp_xlsx):
        spiegel = _spiegel()
        exporter.export(spiegel, tmp_xlsx)
        import os
        assert os.path.exists(tmp_xlsx)

    def test_header(self, exporter, tmp_xlsx):
        spiegel = _spiegel()
        exporter.export(spiegel, tmp_xlsx)
        wb = load_workbook(tmp_xlsx)
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "Testprojekt"
        assert ws.cell(row=2, column=1).value == "Preisspiegel"

    def test_column_headers_two_bidders(self, exporter, tmp_xlsx):
        bidders = [_bidder("Firma A"), _bidder("Firma B")]
        spiegel = _spiegel(bidders=bidders)
        exporter.export(spiegel, tmp_xlsx)
        wb = load_workbook(tmp_xlsx)
        ws = wb.active
        headers = [ws.cell(row=4, column=c).value for c in range(1, 12)]
        assert headers[0] == "OZ"
        assert headers[1] == "Kurztext"
        assert headers[2] == "Menge"
        assert headers[3] == "Einheit"
        assert headers[4] == "Firma A EP"
        assert headers[5] == "Firma A GP"
        assert headers[6] == "Firma B EP"
        assert headers[7] == "Firma B GP"
        assert headers[8] == "Min EP"
        assert headers[9] == "Max EP"
        assert headers[10] == "Avg EP"

    def test_item_row_data(self, exporter, tmp_xlsx):
        bidders = [_bidder("A")]
        row = _item_row(
            unit_prices=[Decimal("5.00")],
            total_prices=[Decimal("50.00")],
            min_up=Decimal("5.00"),
            max_up=Decimal("5.00"),
            avg_up=Decimal("5.00"),
        )
        spiegel = _spiegel(bidders=bidders, rows=[row])
        exporter.export(spiegel, tmp_xlsx)
        wb = load_workbook(tmp_xlsx)
        ws = wb.active
        # Data starts at row 5
        assert ws.cell(row=5, column=1).value == "01.0010"
        assert ws.cell(row=5, column=2).value == "Beton"
        assert ws.cell(row=5, column=3).value == 10.0
        assert ws.cell(row=5, column=4).value == "m2"
        assert ws.cell(row=5, column=5).value == 5.0   # A EP
        assert ws.cell(row=5, column=6).value == 50.0   # A GP

    def test_category_row(self, exporter, tmp_xlsx):
        bidders = [_bidder("A")]
        cat = _cat_row(totals=[Decimal("1000.00")])
        spiegel = _spiegel(bidders=bidders, rows=[cat])
        exporter.export(spiegel, tmp_xlsx)
        wb = load_workbook(tmp_xlsx)
        ws = wb.active
        assert ws.cell(row=5, column=1).value == "01"
        assert ws.cell(row=5, column=2).value == "Rohbau"
        # GP column for bidder 0 = column 6
        assert ws.cell(row=5, column=6).value == 1000.0

    def test_not_offered_shows_na(self, exporter, tmp_xlsx):
        bidders = [_bidder("A")]
        row = _item_row(
            unit_prices=[None],
            total_prices=[None],
            not_offered=[True],
        )
        spiegel = _spiegel(bidders=bidders, rows=[row])
        exporter.export(spiegel, tmp_xlsx)
        wb = load_workbook(tmp_xlsx)
        ws = wb.active
        assert ws.cell(row=5, column=5).value == "n.a."
        assert ws.cell(row=5, column=6).value == "n.a."

    def test_grand_total(self, exporter, tmp_xlsx):
        bidders = [_bidder("A"), _bidder("B")]
        spiegel = _spiegel(
            bidders=bidders,
            grand_totals=[Decimal("5000.00"), Decimal("6000.00")],
        )
        exporter.export(spiegel, tmp_xlsx)
        wb = load_workbook(tmp_xlsx)
        ws = wb.active
        # Grand total row (row 5 since no data rows)
        found = False
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "Gesamtsumme":
                assert ws.cell(row=r, column=6).value == 5000.0   # A GP
                assert ws.cell(row=r, column=8).value == 6000.0   # B GP
                found = True
                break
        assert found

    def test_min_max_highlighting(self, exporter, tmp_xlsx):
        bidders = [_bidder("A"), _bidder("B")]
        row = _item_row(
            unit_prices=[Decimal("5.00"), Decimal("10.00")],
            total_prices=[Decimal("50.00"), Decimal("100.00")],
            not_offered=[False, False],
            min_up=Decimal("5.00"),
            max_up=Decimal("10.00"),
            avg_up=Decimal("7.50"),
        )
        spiegel = _spiegel(bidders=bidders, rows=[row])
        exporter.export(spiegel, tmp_xlsx)
        wb = load_workbook(tmp_xlsx)
        ws = wb.active
        # A EP (col 5) should have MIN fill (green)
        a_ep = ws.cell(row=5, column=5)
        assert a_ep.fill.start_color.rgb == "00C6EFCE"
        # B EP (col 7) should have MAX fill (red)
        b_ep = ws.cell(row=5, column=7)
        assert b_ep.fill.start_color.rgb == "00FFC7CE"

    def test_empty_spiegel(self, exporter, tmp_xlsx):
        spiegel = _spiegel()
        exporter.export(spiegel, tmp_xlsx)
        wb = load_workbook(tmp_xlsx)
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "Testprojekt"
