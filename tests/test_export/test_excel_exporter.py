import os
import tempfile
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from lvgenerator.constants import GAEBPhase
from lvgenerator.export.excel_exporter import ExcelExporter
from lvgenerator.models.boq import BoQ, BoQInfo
from lvgenerator.models.category import BoQCategory
from lvgenerator.models.item import Item, ItemDescription
from lvgenerator.models.project import GAEBProject, PrjInfo, AwardInfo, GAEBInfo


def _make_project(phase: GAEBPhase, categories=None):
    return GAEBProject(
        gaeb_info=GAEBInfo(),
        prj_info=PrjInfo(name="Testprojekt"),
        award_info=AwardInfo(),
        phase=phase,
        boq=BoQ(
            id="boq-1", info=BoQInfo(name="Test-LV"),
            categories=categories or [],
        ),
    )


def _make_item(id, rno, qty=None, qu="", up=None, it=None, text="Position"):
    return Item(
        id=id, rno_part=rno, qty=qty, qu=qu, up=up, it=it,
        description=ItemDescription(outline_text=text, detail_text=f"Langtext {rno}"),
    )


@pytest.fixture
def exporter():
    return ExcelExporter()


@pytest.fixture
def tmp_xlsx(tmp_path):
    return str(tmp_path / "test_output.xlsx")


class TestExcelExporter:
    def test_export_creates_file(self, exporter, tmp_xlsx):
        project = _make_project(GAEBPhase.X83)
        exporter.export(project, tmp_xlsx)
        assert os.path.exists(tmp_xlsx)

    def test_project_name_in_header(self, exporter, tmp_xlsx):
        project = _make_project(GAEBPhase.X83)
        exporter.export(project, tmp_xlsx)
        wb = load_workbook(tmp_xlsx)
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "Testprojekt"

    def test_phase_in_header(self, exporter, tmp_xlsx):
        project = _make_project(GAEBPhase.X84)
        exporter.export(project, tmp_xlsx)
        wb = load_workbook(tmp_xlsx)
        ws = wb.active
        assert "X84" in ws.cell(row=2, column=1).value

    def test_x83_no_price_columns(self, exporter, tmp_xlsx):
        cat = BoQCategory(id="c1", rno_part="01", label="Rohbau", items=[
            _make_item("i1", "0010", qty=Decimal("10"), qu="m2"),
        ])
        project = _make_project(GAEBPhase.X83, [cat])
        exporter.export(project, tmp_xlsx)
        wb = load_workbook(tmp_xlsx)
        ws = wb.active
        headers = [ws.cell(row=4, column=c).value for c in range(1, 10)]
        assert "EP" not in headers
        assert "GP" not in headers
        assert "Menge" in headers

    def test_x84_has_price_columns(self, exporter, tmp_xlsx):
        cat = BoQCategory(id="c1", rno_part="01", label="Rohbau", items=[
            _make_item("i1", "0010", qty=Decimal("10"), qu="m2",
                       up=Decimal("5.00"), it=Decimal("50.00")),
        ])
        project = _make_project(GAEBPhase.X84, [cat])
        exporter.export(project, tmp_xlsx)
        wb = load_workbook(tmp_xlsx)
        ws = wb.active
        headers = [ws.cell(row=4, column=c).value for c in range(1, 10)]
        assert "EP" in headers
        assert "GP" in headers

    def test_category_row_present(self, exporter, tmp_xlsx):
        cat = BoQCategory(id="c1", rno_part="01", label="Betonarbeiten", items=[
            _make_item("i1", "0010", qty=Decimal("10"), qu="m2"),
        ])
        project = _make_project(GAEBPhase.X83, [cat])
        exporter.export(project, tmp_xlsx)
        wb = load_workbook(tmp_xlsx)
        ws = wb.active
        # Category row is row 5 (after header rows 1-4)
        assert ws.cell(row=5, column=1).value == "01"
        assert ws.cell(row=5, column=2).value == "Betonarbeiten"

    def test_item_data_correct(self, exporter, tmp_xlsx):
        cat = BoQCategory(id="c1", rno_part="01", label="Test", items=[
            _make_item("i1", "0010", qty=Decimal("100.000"),
                       qu="m2", up=Decimal("25.00"), it=Decimal("2500.00"),
                       text="Beton C25"),
        ])
        project = _make_project(GAEBPhase.X84, [cat])
        exporter.export(project, tmp_xlsx)
        wb = load_workbook(tmp_xlsx)
        ws = wb.active
        # Item is in row 6 (header 1-3, colheaders 4, category 5, item 6)
        assert ws.cell(row=6, column=1).value == "01.0010"
        assert ws.cell(row=6, column=2).value == "Beton C25"
        assert ws.cell(row=6, column=4).value == 100.0
        assert ws.cell(row=6, column=5).value == "m2"
        assert ws.cell(row=6, column=6).value == 25.0
        assert ws.cell(row=6, column=7).value == 2500.0

    def test_grand_total(self, exporter, tmp_xlsx):
        cat = BoQCategory(id="c1", rno_part="01", label="Test", items=[
            _make_item("i1", "0010", it=Decimal("100.00")),
            _make_item("i2", "0020", it=Decimal("200.00")),
        ])
        project = _make_project(GAEBPhase.X84, [cat])
        exporter.export(project, tmp_xlsx)
        wb = load_workbook(tmp_xlsx)
        ws = wb.active
        # Find grand total row
        found = False
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == "Gesamtsumme":
                assert ws.cell(row=row, column=7).value == 300.0
                found = True
                break
        assert found

    def test_empty_project(self, exporter, tmp_xlsx):
        project = _make_project(GAEBPhase.X83)
        exporter.export(project, tmp_xlsx)
        wb = load_workbook(tmp_xlsx)
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "Testprojekt"

    def test_nested_categories(self, exporter, tmp_xlsx):
        sub = BoQCategory(id="s1", rno_part="01", label="Sub", items=[
            _make_item("i1", "0010"),
        ])
        cat = BoQCategory(
            id="c1", rno_part="01", label="Parent",
            subcategories=[sub], items=[],
        )
        project = _make_project(GAEBPhase.X83, [cat])
        exporter.export(project, tmp_xlsx)
        wb = load_workbook(tmp_xlsx)
        ws = wb.active
        # Should have parent category row and sub category row
        values = [ws.cell(row=r, column=2).value for r in range(5, 9)]
        assert "Parent" in values
        assert "Sub" in values
