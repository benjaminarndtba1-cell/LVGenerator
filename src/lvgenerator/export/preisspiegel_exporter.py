from decimal import Decimal
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from lvgenerator.models.preisspiegel import (
    PreisSpiegel,
    PreisSpiegelCategoryRow,
    PreisSpiegelRow,
)


class PreisSpiegelExporter:
    """Exportiert einen Preisspiegel als Excel-Datei (.xlsx)."""

    HEADER_FONT = Font(bold=True, size=12)
    SUB_HEADER_FONT = Font(bold=True, size=11)
    COL_HEADER_FONT = Font(bold=True, size=10)
    COL_HEADER_FILL = PatternFill(start_color="4472C4", fill_type="solid")
    COL_HEADER_FONT_WHITE = Font(bold=True, size=10, color="FFFFFF")
    CATEGORY_FONT = Font(bold=True, size=10)
    CATEGORY_FILL = PatternFill(start_color="D9E1F2", fill_type="solid")
    TOTAL_FONT = Font(bold=True, size=10)
    MIN_FILL = PatternFill(start_color="C6EFCE", fill_type="solid")
    MAX_FILL = PatternFill(start_color="FFC7CE", fill_type="solid")
    NUMBER_FORMAT = '#,##0.00'
    QTY_FORMAT = '#,##0.000'

    def export(self, spiegel: PreisSpiegel, file_path: str) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Preisspiegel"

        n = len(spiegel.bidders)

        # Header
        ws.cell(row=1, column=1, value=spiegel.project_name).font = self.HEADER_FONT
        ws.cell(row=2, column=1, value="Preisspiegel").font = self.SUB_HEADER_FONT

        # Column headers
        headers = self._get_headers(spiegel)
        row = 4
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.COL_HEADER_FONT_WHITE
            cell.fill = self.COL_HEADER_FILL
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        row += 1

        # Data rows
        for data_row in spiegel.rows:
            if isinstance(data_row, PreisSpiegelCategoryRow):
                row = self._write_category_row(ws, data_row, headers, row, n)
            else:
                row = self._write_item_row(ws, data_row, headers, row, n)

        # Grand total
        if spiegel.grand_totals:
            self._write_grand_total(ws, spiegel, headers, row, n)

        self._auto_fit_columns(ws, headers, n)
        wb.save(file_path)

    def _get_headers(self, spiegel: PreisSpiegel) -> list[str]:
        headers = ["OZ", "Kurztext", "Menge", "Einheit"]
        for bidder in spiegel.bidders:
            headers.append(f"{bidder.name} EP")
            headers.append(f"{bidder.name} GP")
        headers.extend(["Min EP", "Max EP", "Avg EP"])
        return headers

    def _write_category_row(
        self, ws, cat_row: PreisSpiegelCategoryRow,
        headers: list[str], row: int, n: int,
    ) -> int:
        total_cols = len(headers)
        ws.cell(row=row, column=1, value=cat_row.oz).font = self.CATEGORY_FONT
        ws.cell(row=row, column=2, value=cat_row.label).font = self.CATEGORY_FONT
        for col in range(1, total_cols + 1):
            ws.cell(row=row, column=col).fill = self.CATEGORY_FILL

        # Category totals (GP columns only)
        for i, total in enumerate(cat_row.totals):
            if total is not None:
                gp_col = 4 + i * 2 + 2  # GP column for bidder i
                cell = ws.cell(row=row, column=gp_col, value=float(total))
                cell.number_format = self.NUMBER_FORMAT
                cell.font = self.TOTAL_FONT
                cell.fill = self.CATEGORY_FILL

        return row + 1

    def _write_item_row(
        self, ws, item_row: PreisSpiegelRow,
        headers: list[str], row: int, n: int,
    ) -> int:
        col = 1
        ws.cell(row=row, column=col, value=item_row.oz)
        col += 1
        ws.cell(row=row, column=col, value=item_row.short_text)
        col += 1

        if item_row.qty is not None:
            c = ws.cell(row=row, column=col, value=float(item_row.qty))
            c.number_format = self.QTY_FORMAT
        col += 1
        ws.cell(row=row, column=col, value=item_row.qu)
        col += 1

        # Bidder EP/GP pairs
        for i in range(n):
            # EP
            if item_row.not_offered[i]:
                ws.cell(row=row, column=col, value="n.a.")
            elif item_row.unit_prices[i] is not None:
                c = ws.cell(row=row, column=col, value=float(item_row.unit_prices[i]))
                c.number_format = self.NUMBER_FORMAT
                # Highlight min/max
                if item_row.min_up is not None and item_row.unit_prices[i] == item_row.min_up and n > 1:
                    c.fill = self.MIN_FILL
                if item_row.max_up is not None and item_row.unit_prices[i] == item_row.max_up and n > 1:
                    c.fill = self.MAX_FILL
            col += 1

            # GP
            if item_row.not_offered[i]:
                ws.cell(row=row, column=col, value="n.a.")
            elif item_row.total_prices[i] is not None:
                c = ws.cell(row=row, column=col, value=float(item_row.total_prices[i]))
                c.number_format = self.NUMBER_FORMAT
            col += 1

        # Statistics
        if item_row.min_up is not None:
            c = ws.cell(row=row, column=col, value=float(item_row.min_up))
            c.number_format = self.NUMBER_FORMAT
        col += 1
        if item_row.max_up is not None:
            c = ws.cell(row=row, column=col, value=float(item_row.max_up))
            c.number_format = self.NUMBER_FORMAT
        col += 1
        if item_row.avg_up is not None:
            c = ws.cell(row=row, column=col, value=float(item_row.avg_up))
            c.number_format = self.NUMBER_FORMAT

        return row + 1

    def _write_grand_total(
        self, ws, spiegel: PreisSpiegel,
        headers: list[str], row: int, n: int,
    ) -> None:
        total_cols = len(headers)
        ws.cell(row=row, column=1, value="Gesamtsumme").font = self.TOTAL_FONT

        for i, total in enumerate(spiegel.grand_totals):
            if total is not None:
                gp_col = 4 + i * 2 + 2  # GP column for bidder i
                cell = ws.cell(row=row, column=gp_col, value=float(total))
                cell.number_format = self.NUMBER_FORMAT
                cell.font = self.TOTAL_FONT

    def _auto_fit_columns(self, ws, headers: list[str], n: int) -> None:
        widths = {"OZ": 14, "Kurztext": 30, "Menge": 12, "Einheit": 10}
        for col, header in enumerate(headers, 1):
            if header in widths:
                width = widths[header]
            elif header.endswith(" EP") or header.endswith(" GP"):
                width = 14
            elif header in ("Min EP", "Max EP", "Avg EP"):
                width = 12
            else:
                width = 14
            ws.column_dimensions[get_column_letter(col)].width = width
