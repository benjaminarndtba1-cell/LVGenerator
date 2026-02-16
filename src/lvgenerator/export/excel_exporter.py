from decimal import Decimal
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from lvgenerator.constants import GAEBPhase
from lvgenerator.gaeb.phase_rules import get_rules
from lvgenerator.models.category import BoQCategory
from lvgenerator.models.item import Item
from lvgenerator.models.project import GAEBProject


class ExcelExporter:
    """Exportiert ein GAEB-Projekt als Excel-Datei (.xlsx)."""

    HEADER_FONT = Font(bold=True, size=12)
    CATEGORY_FONT = Font(bold=True, size=11)
    CATEGORY_FILL = PatternFill(start_color="D9E1F2", fill_type="solid")
    COL_HEADER_FONT = Font(bold=True, size=10)
    TOTAL_FONT = Font(bold=True, size=10)
    NUMBER_FORMAT = '#,##0.00'
    QTY_FORMAT = '#,##0.000'

    def export(self, project: GAEBProject, file_path: str) -> None:
        """Exportiert das Projekt nach Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Leistungsverzeichnis"

        rules = get_rules(project.phase)
        headers = self._get_headers(rules)

        # Projektkopf
        ws.cell(row=1, column=1, value=project.prj_info.name).font = self.HEADER_FONT
        ws.cell(row=2, column=1,
                value=f"Phase: {project.phase.name} - {project.phase.label_de}")

        # Spaltenüberschriften
        row = 4
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.COL_HEADER_FONT
        row += 1

        # Daten
        if project.boq:
            row = self._write_categories(
                ws, project.boq.categories, rules, headers, row, ""
            )

            # Gesamtsumme
            if rules.has_totals and project.boq.categories:
                grand_total = Decimal("0.00")
                has_any = False
                for cat in project.boq.categories:
                    t = cat.calculate_total()
                    if t is not None:
                        grand_total += t
                        has_any = True
                if has_any:
                    gp_col = len(headers)
                    ws.cell(row=row, column=1, value="Gesamtsumme").font = self.TOTAL_FONT
                    cell = ws.cell(row=row, column=gp_col, value=float(grand_total))
                    cell.number_format = self.NUMBER_FORMAT
                    cell.font = self.TOTAL_FONT

        self._auto_fit_columns(ws, headers)
        wb.save(file_path)

    def _get_headers(self, rules) -> list[str]:
        headers = ["OZ", "Kurztext", "Langtext"]
        if rules.has_quantities:
            headers.extend(["Menge", "Einheit"])
        if rules.has_prices:
            headers.append("EP")
        if rules.has_totals:
            headers.append("GP")
        return headers

    def _write_categories(self, ws, categories: list[BoQCategory],
                          rules, headers: list[str],
                          row: int, parent_oz: str) -> int:
        for cat in categories:
            oz = f"{parent_oz}.{cat.rno_part}" if parent_oz else cat.rno_part
            # Kategoriezeile
            ws.cell(row=row, column=1, value=oz).font = self.CATEGORY_FONT
            ws.cell(row=row, column=2, value=cat.label).font = self.CATEGORY_FONT
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = self.CATEGORY_FILL

            if rules.has_totals:
                total = cat.calculate_total()
                if total is not None:
                    gp_col = len(headers)
                    cell = ws.cell(row=row, column=gp_col, value=float(total))
                    cell.number_format = self.NUMBER_FORMAT
                    cell.font = self.TOTAL_FONT
                    cell.fill = self.CATEGORY_FILL
            row += 1

            # Unterkategorien
            row = self._write_categories(
                ws, cat.subcategories, rules, headers, row, oz
            )

            # Positionen
            for item in cat.items:
                row = self._write_item(ws, item, rules, headers, row, oz)

        return row

    def _write_item(self, ws, item: Item, rules, headers: list[str],
                    row: int, parent_oz: str) -> int:
        col = 1
        full_oz = f"{parent_oz}.{item.rno_part}" if parent_oz else item.rno_part
        ws.cell(row=row, column=col, value=full_oz)
        col += 1
        ws.cell(row=row, column=col, value=item.description.outline_text)
        col += 1
        detail = item.description.detail_text
        cell = ws.cell(row=row, column=col, value=detail)
        cell.alignment = Alignment(wrap_text=True)
        col += 1

        if rules.has_quantities:
            if item.qty is not None:
                c = ws.cell(row=row, column=col, value=float(item.qty))
                c.number_format = self.QTY_FORMAT
            col += 1
            ws.cell(row=row, column=col, value=item.qu)
            col += 1

        if rules.has_prices:
            if item.up is not None:
                c = ws.cell(row=row, column=col, value=float(item.up))
                c.number_format = self.NUMBER_FORMAT
            col += 1

        if rules.has_totals:
            total = item.it if item.it is not None else item.calculate_total()
            if total is not None:
                c = ws.cell(row=row, column=col, value=float(total))
                c.number_format = self.NUMBER_FORMAT

        return row + 1

    def _auto_fit_columns(self, ws, headers: list[str]) -> None:
        min_widths = {"OZ": 12, "Kurztext": 30, "Langtext": 40, "Menge": 12,
                      "Einheit": 10, "EP": 14, "GP": 14}
        for col, header in enumerate(headers, 1):
            width = min_widths.get(header, 12)
            ws.column_dimensions[get_column_letter(col)].width = width
